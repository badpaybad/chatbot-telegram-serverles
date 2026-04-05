import unittest
import os
import sys
import torch
import numpy as np
from PIL import Image

# Import current module
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.append(project_root)

from gemma4 import describe_image, query_image, process_file_with_prompt, read_file_content

class TestGemma4Update3(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Paths for testing
        cls.image_file = os.path.join(project_root, "img1.png")
        cls.pdf_file = os.path.join(project_root, "test", "dummy.pdf")
        cls.txt_file = os.path.join(project_root, "test", "dummy.txt")
        cls.docx_file = os.path.join(project_root, "test", "dummy.docx")
        cls.xlsx_file = os.path.join(project_root, "test", "dummy.xlsx")

        # Create dummy files if they don't exist for parsing verification
        os.makedirs(os.path.join(project_root, "test"), exist_ok=True)
        
        with open(cls.txt_file, "w", encoding="utf-8") as f:
            f.write("Đây là nội dung thử nghiệm của file TXT.")

        # For PDF/DOCX/XLSX, we need to use libraries to create them if we want to test extraction
        # Since we've verified libraries are present, let's create a minimal docx/xlsx if possible
        try:
            import docx
            doc = docx.Document()
            doc.add_paragraph("Nội dung thử nghiệm của file DOCX.")
            doc.save(cls.docx_file)
            
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "Họ tên"
            ws["B1"] = "Tuổi"
            ws["A2"] = "Nguyen Van A"
            ws["B2"] = 30
            wb.save(cls.xlsx_file)
        except Exception as e:
            print(f"Warning: Could not create dummy docx/xlsx: {e}")

    def test_image_description(self):
        print("\n[*] Testing Image Description...")
        if os.path.exists(self.image_file):
            desc = describe_image(self.image_file)
            print(f"Description: {desc}")
            self.assertIsInstance(desc, str)
            self.assertTrue(len(desc) > 0)
        else:
            print("[!] Skipping Image Description test: img1.png not found")

    def test_image_query(self):
        print("\n[*] Testing Image Query (VQA)...")
        if os.path.exists(self.image_file):
            ans = query_image(self.image_file, "Trong ảnh có gì nổi bật?")
            print(f"Answer: {ans}")
            self.assertIsInstance(ans, str)
        else:
            print("[!] Skipping Image Query test: img1.png not found")

    def test_file_reading_txt(self):
        print("\n[*] Testing TXT reading...")
        content = read_file_content(self.txt_file)
        self.assertIn("nội dung thử nghiệm", content)
        print(f"Content: {content}")

    def test_file_reading_docx(self):
        print("\n[*] Testing DOCX reading...")
        if os.path.exists(self.docx_file):
            content = read_file_content(self.docx_file)
            self.assertIn("Nội dung thử nghiệm", content)
            print(f"Content: {content}")

    def test_file_reading_xlsx(self):
        print("\n[*] Testing XLSX reading...")
        if os.path.exists(self.xlsx_file):
            content = read_file_content(self.xlsx_file)
            self.assertIn("Nguyen Van A", content)
            print(f"Content: {content}")

    def test_process_file_with_prompt(self):
        print("\n[*] Testing process_file_with_prompt (LLM integration)...")
        result = process_file_with_prompt(self.txt_file, "Tóm tắt file này.")
        print(f"LLM Result: {result}")
        self.assertIsInstance(result, str)

if __name__ == "__main__":
    # Remove config_dunp from sys.argv so unittest doesn't complain
    sys.argv.pop(1)
    unittest.main()
